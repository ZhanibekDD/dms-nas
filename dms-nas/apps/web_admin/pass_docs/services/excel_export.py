"""
Генерация Excel-заявки АСДПО для сотрудника по шаблону.

Шаблон: pass_docs/templates_excel/asdpo_template.xlsx
Лист «Реестр» — данные сотрудника начинаются со строки 5.

Маппинг колонок (по шаблону АСДПО_ЗаявкаСотрудники):
  E  — Фамилия           F  — Имя              G  — Отчество
  H  — Дата рождения     I  — Гражданство       J  — Тип документа
  K  — Серия паспорта    L  — Номер паспорта    M  — Дата выдачи
  N  — Кем выдан         T  — Профессия/должность
  X  — Лицензия мед. организации               Y  — Медкомиссия (до)
  AB — Охрана труда (А,Б 3 года)               AF — Номер протокола ЕИСОТ
  AE — СИЗ (3 года)      AI — ПБ А.1 (5 лет)   AJ — ПБ Б.2.1
  AN — Нефтепромысловые трубопроводы           AW — ПТМ (5 лет)
  AZ — Электробезопасность (1 год)             BN — БДД (1 год)
  BQ — Дата трудового договора                 BS — № сертификата ПБ1
  BT — Дата ПБ1
"""
from __future__ import annotations

import io
import logging
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

TEMPLATE_PATH = Path(__file__).resolve().parent.parent / "templates_excel" / "asdpo_template.xlsx"

# document_type.code (нормализованный) → {col: поле в normalized}
# Коды числовые — как в каталоге pass_docs/catalog/document_codes.py
_CODE_COLS: dict[str, dict[str, str]] = {
    "13":  {"AA": "issue_date"},                           # ОТ программа В, 1 год
    "14":  {"AB": "issue_date", "AF": "protocol_number"},  # ОТ А,Б, 3 года + ЕИСОТ
    "15":  {"AW": "issue_date"},                           # ПТМ / пожарная безопасность
    "16":  {},                                             # Газобезопасность (нет чёткой колонки)
    "17":  {"AI": "issue_date"},                           # ПБ общие А.1, 5 лет
    "18":  {"AB": "issue_date"},                           # ОТ доп. протокол
    "19":  {},                                             # ОПП
    "20":  {"AS": "issue_date"},                           # Подъёмные сооружения Б.9.3
    "21":  {"AJ": "issue_date"},                           # ПБ протокол Б.2.1
    "22":  {"AB": "issue_date"},                           # Обучение по ОТ
    "26":  {"AZ": "issue_date"},                           # Электробезопасность
    "31":  {"BN": "issue_date"},                           # БДД
    "44":  {"BQ": "issue_date"},                           # Трудовой договор
    "57":  {"AE": "issue_date"},                           # СИЗ, 3 года
    "59":  {"AW": "issue_date"},                           # ПТМ (второй вариант кода)
    "61":  {"BS": "protocol_number", "BT": "issue_date"},  # ПБ1 — № сертификата + дата
    "74":  {},                                             # УМО
    "78":  {"AN": "issue_date"},                           # Нефтепромысловые трубопроводы
}

# extractor_kind → {col: поле в normalized}  (используется как fallback)
_EKIND_COLS: dict[str, dict[str, str]] = {
    "ru_passport": {
        "K": "series",
        "L": "number",
        "M": "issue_date",
        "N": "issuer",
    },
    "medical_certificate": {
        "X": "certificate_number",
        "Y": "valid_until",
    },
    "electrical_safety": {"AZ": "issue_date"},
    "bdd_protocol":       {"BN": "issue_date"},
    "siz_training_protocol": {"AE": "issue_date"},
}

# Текстовые коды файлов → числовой код каталога
_TEXT_TO_NUM: dict[str, str] = {
    "PASSPORT_RF": "6", "PASPORT_RF": "6",
    "MED": "7", "MEDICAL": "7", "MEDICAL_CERTIFICATE": "7",
    "MED_SPRAVKA": "7", "SPRAVKA086": "7",
}


def _norm_code(raw: str) -> str:
    c = (raw or "").strip().upper()
    return _TEXT_TO_NUM.get(c, c.lstrip("0") if c.isdigit() else c)


def _as_date(val: Any) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, date):
        return val
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _get_normalized(doc: Any) -> dict:
    ej = (doc.extracted_json or {}) if doc else {}
    return ej.get("normalized") or {}


def _apply_mapping(row: dict, cols_map: dict[str, str], normalized: dict) -> None:
    for col, field in cols_map.items():
        val = normalized.get(field)
        if not val:
            continue
        if "date" in field:
            parsed = _as_date(val)
            if parsed:
                row.setdefault(col, parsed)
        else:
            row.setdefault(col, str(val).strip())


def collect_employee_row(emp: Any) -> dict[str, Any]:
    """Собрать данные одного сотрудника для строки Реестра."""
    from pass_docs.models import EmployeeDocument

    row: dict[str, Any] = {
        "E": emp.last_name or "",
        "F": emp.first_name or "",
        "G": emp.middle_name or "",
        "H": emp.birth_date,
        "I": "643",
        "J": "паспорт гражданина РФ",
        "K": emp.passport_series or "",
        "L": emp.passport_number or "",
        "T": emp.profession_label or "",
    }

    docs = (
        EmployeeDocument.objects
        .filter(employee=emp, is_actual=True)
        .select_related("document_type")
        .order_by("document_type__sort_order", "-updated_at")
    )

    for doc in docs:
        code = _norm_code(doc.document_type.code or "")
        ek = (doc.document_type.extractor_kind or "").strip()
        normalized = _get_normalized(doc)

        if code in _CODE_COLS:
            _apply_mapping(row, _CODE_COLS[code], normalized)

        if ek in _EKIND_COLS:
            _apply_mapping(row, _EKIND_COLS[ek], normalized)

    # Перекрыть manual_data (ручные правки имеют приоритет)
    manual = emp.manual_data or {}
    _manual_map = {
        "last_name":          "E",
        "first_name":         "F",
        "middle_name":        "G",
        "birth_date":         "H",
        "profession_label":   "T",
        "passport_series":    "K",
        "passport_number":    "L",
        "passport_issue_date": "M",
        "passport_issuer":    "N",
    }
    for mkey, col in _manual_map.items():
        val = manual.get(mkey)
        if not val:
            continue
        if "date" in mkey:
            parsed = _as_date(val)
            if parsed:
                row[col] = parsed
        else:
            row[col] = str(val).strip()

    # Очистить пустые строки
    return {k: v for k, v in row.items() if v not in (None, "")}


def _copy_row_style(ws_src: Any, ws_dst: Any, src_row: int, dst_row: int) -> None:
    """Перенести стили строки шаблона на новую строку."""
    from openpyxl.utils import get_column_letter
    max_col = ws_src.max_column
    for col in range(1, max_col + 1):
        src = ws_src.cell(row=src_row, column=col)
        dst = ws_dst.cell(row=dst_row, column=col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.number_format = src.number_format
            dst.alignment = copy(src.alignment)


def generate_excel(employees: list[Any]) -> bytes:
    """
    Вернуть байты .xlsx с заполненным листом «Реестр».
    employees — список объектов Employee (один или несколько).
    """
    import openpyxl
    from openpyxl.utils import column_index_from_string

    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"Шаблон Excel не найден: {TEMPLATE_PATH}. "
            "Скопируйте файл АСДПО_Шаблон_ЗаявкаСотрудники.xlsx в "
            "pass_docs/templates_excel/asdpo_template.xlsx"
        )

    wb = openpyxl.load_workbook(str(TEMPLATE_PATH))
    ws = wb["Реестр"]

    DATA_ROW_START = 5  # первая строка данных в шаблоне

    for idx, emp in enumerate(employees):
        dst_row = DATA_ROW_START + idx
        if idx > 0:
            _copy_row_style(ws, ws, DATA_ROW_START, dst_row)

        row_data = collect_employee_row(emp)
        row_data["A"] = idx + 1  # №п/п

        for col_letter, value in row_data.items():
            try:
                col_idx = column_index_from_string(col_letter)
            except Exception:
                continue
            cell = ws.cell(row=dst_row, column=col_idx)
            cell.value = value

        logger.debug("Excel row %d: %s — %d fields", dst_row, emp.full_name, len(row_data))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def generate_excel_one(emp: Any) -> bytes:
    return generate_excel([emp])
