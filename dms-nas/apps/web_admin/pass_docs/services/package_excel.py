"""
Генерация операторского XLSX для PackageRequest (листы summary, validation, package_meta).
"""

from __future__ import annotations

import json
from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font

from pass_docs.models import EmployeeDocument, PackageRequest
from pass_docs.services.package_validation import PackageDocumentResolution, ValidationEntry, extractor_kind_for_document


def _norm_block(doc: EmployeeDocument) -> dict[str, Any]:
    payload = doc.extracted_json if isinstance(doc.extracted_json, dict) else {}
    n = payload.get("normalized")
    return n if isinstance(n, dict) else {}


def format_normalized_main_fields(doc: EmployeeDocument) -> str:
    """
    Короткая строка по типу экстрактора для листа summary.
    """
    norm = _norm_block(doc)
    schema = (norm.get("schema") or "").strip().lower()
    kind = extractor_kind_for_document(doc).lower()

    def trunc(s: str, n: int = 400) -> str:
        t = (s or "").replace("\n", " ").strip()
        return t if len(t) <= n else t[: n - 1] + "…"

    if schema == "ru_passport" or kind == "ru_passport":
        series = norm.get("series") or ""
        number = norm.get("number") or ""
        fn = norm.get("full_number") or ""
        pair = f"{series} {number}".strip() or fn
        bd = norm.get("birth_date") or ""
        parts = [f"серия+номер: {pair}".strip(), f"дата рождения: {bd}".strip()]
        return trunc("; ".join(p for p in parts if p and not p.endswith(": ")))

    if schema == "medical_certificate" or kind == "medical_certificate":
        org = norm.get("organization") or ""
        issue = norm.get("issue_date") or ""
        concl = norm.get("conclusion") or ""
        parts = [
            f"организация: {org}".strip() if org else "",
            f"дата: {issue}".strip() if issue else "",
            f"conclusion: {concl}".strip() if concl else "",
        ]
        return trunc("; ".join(p for p in parts if p))

    training_kinds = frozenset(
        {
            "safety_protocol_v",
            "safety_protocol_ab",
            "electrical_safety",
            "bdd_protocol",
            "siz_training_protocol",
            "umo",
        }
    )
    if kind in training_kinds or schema in training_kinds:
        pn = norm.get("protocol_number") or ""
        issue = norm.get("issue_date") or ""
        prog = norm.get("program_name") or ""
        concl = norm.get("conclusion") or ""
        parts = [
            f"номер: {pn}".strip() if pn else "",
            f"дата: {issue}".strip() if issue else "",
            f"программа: {prog}".strip() if prog else "",
            f"conclusion: {concl}".strip() if concl else "",
        ]
        return trunc("; ".join(p for p in parts if p))

    # Прочие / неизвестные: компактный JSON фрагмент
    try:
        return trunc(json.dumps(norm, ensure_ascii=False)[:350])
    except Exception:
        return ""


def _issue_expiry_str(doc: EmployeeDocument) -> tuple[str, str]:
    issue = doc.issue_date.isoformat() if doc.issue_date else ""
    exp = doc.expiry_date.isoformat() if doc.expiry_date else ""
    return issue, exp


def build_package_workbook(
    request: PackageRequest,
    resolution: PackageDocumentResolution,
    *,
    built_at: datetime | None = None,
) -> bytes:
    wb = Workbook()
    # --- summary ---
    ws1 = wb.active
    ws1.title = "summary"
    headers = [
        "employee_import_key",
        "employee_full_name",
        "document_id",
        "doc_code",
        "doc_name",
        "extractor_kind",
        "parse_status",
        "is_actual",
        "issue_date",
        "expiry_date",
        "source_path",
        "normalized_main_fields",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)

    emp = resolution.employee
    row = 2
    for doc in resolution.included:
        issue_s, exp_s = _issue_expiry_str(doc)
        ws1.cell(row=row, column=1, value=emp.import_key)
        ws1.cell(row=row, column=2, value=emp.full_name)
        ws1.cell(row=row, column=3, value=doc.pk)
        ws1.cell(row=row, column=4, value=doc.document_type.code)
        ws1.cell(row=row, column=5, value=doc.document_type.name)
        ws1.cell(row=row, column=6, value=extractor_kind_for_document(doc))
        ws1.cell(row=row, column=7, value=doc.parse_status)
        ws1.cell(row=row, column=8, value="yes" if doc.is_actual else "no")
        ws1.cell(row=row, column=9, value=issue_s)
        ws1.cell(row=row, column=10, value=exp_s)
        ws1.cell(row=row, column=11, value=doc.source_path)
        ws1.cell(row=row, column=12, value=format_normalized_main_fields(doc))
        row += 1

    # --- validation ---
    ws2 = wb.create_sheet("validation")
    v_headers = ["document_id", "doc_code", "severity", "message", "field", "current_value"]
    for col, h in enumerate(v_headers, start=1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
    r = 2
    for e in resolution.validation_entries:
        ws2.cell(row=r, column=1, value=e.document_id)
        ws2.cell(row=r, column=2, value=e.doc_code)
        ws2.cell(row=r, column=3, value=e.severity)
        ws2.cell(row=r, column=4, value=e.message)
        ws2.cell(row=r, column=5, value=e.field)
        ws2.cell(row=r, column=6, value=e.current_value)
        r += 1

    # --- package_meta ---
    ws3 = wb.create_sheet("package_meta")
    ws3.cell(row=1, column=1, value="key")
    ws3.cell(row=1, column=2, value="value")
    ws3["A1"].font = Font(bold=True)
    ws3["B1"].font = Font(bold=True)

    bt = built_at or datetime.now(timezone.utc)
    if bt.tzinfo is None:
        bt = bt.replace(tzinfo=timezone.utc)
    built_iso = bt.isoformat()

    payload = request.payload_json if isinstance(request.payload_json, dict) else {}
    filters = payload.get("filters") if isinstance(payload.get("filters"), dict) else {}

    meta_rows: list[tuple[str, Any]] = [
        ("request_id", request.pk),
        ("employee_id", request.employee_id),
        ("employee_import_key", request.employee.import_key),
        ("employee_full_name", request.employee.full_name),
        ("built_at", built_iso),
        ("documents_total", resolution.documents_total),
        ("documents_included", resolution.documents_included),
        ("filters_only_actual", filters.get("only_actual", "")),
        ("filters_only_parse_ok", filters.get("only_parse_ok", "")),
        ("package_kind", request.package_kind or payload.get("package_kind") or ""),
        ("status_after_build", request.status),
        ("build_notes", payload.get("build_notes") or ""),
    ]
    for i, (k, v) in enumerate(meta_rows, start=2):
        ws3.cell(row=i, column=1, value=k)
        ws3.cell(row=i, column=2, value=str(v) if v is not None else "")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
