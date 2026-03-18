"""
Django admin registrations with custom actions and filters.
Sprint 10/11: Document Registry, UserObject, mass operations.
"""

from django.contrib import admin
from django.http import HttpResponse
from django.utils.html import format_html
import csv
import io
import sys
import os

from .models import (
    BotUser, UploadLog, ExpiryItem, FinanceDoc,
    FinanceStatusLog, Problem, PackageLog, Report, AuditLog,
    DocLink, NasObject, Document, UserObject, OcrResult,
)

# ── Helpers ───────────────────────────────────────────────────────────────────

def _ensure_root_in_path():
    root = os.path.dirname(
        os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    )
    if root not in sys.path:
        sys.path.insert(0, root)


def _nas():
    """Get NAS client instance (lazy import)."""
    _ensure_root_in_path()
    from apps.bot.bot_nas import get_nas
    return get_nas()


# ── Permissions by object — mixin (Sprint 9.1+) ───────────────────────────────

class ObjectFilterMixin:
    """
    Фильтрует queryset по разрешённым объектам.
    Суперпользователи и пользователи без записей user_objects видят всё.
    Обычные пользователи: username должен быть их Telegram ID.
    """
    object_field = "object_name"

    def _get_allowed_objects(self, request):
        if request.user.is_superuser:
            return None
        _ensure_root_in_path()
        import apps.bot.bot_db as db
        try:
            tg_id = int(request.user.username)
        except (ValueError, TypeError):
            return None
        allowed = db.get_allowed_objects(tg_id)
        return allowed if allowed else None

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        allowed = self._get_allowed_objects(request)
        if allowed is not None:
            qs = qs.filter(**{f"{self.object_field}__in": allowed})
        return qs


# ── BotUser ───────────────────────────────────────────────────────────────────

@admin.register(BotUser)
class BotUserAdmin(admin.ModelAdmin):
    list_display   = ("telegram_id", "full_name", "username", "role_badge", "is_active", "created_at")
    list_filter    = ("role", "is_active")
    search_fields  = ("full_name", "username", "telegram_id")
    list_editable  = ("is_active",)
    readonly_fields = ("telegram_id", "created_at")

    @admin.display(description="Роль")
    def role_badge(self, obj):
        colors = {
            "admin": "#dc3545", "pto": "#0d6efd", "tb": "#198754",
            "buh": "#fd7e14", "prorab": "#6610f2", "viewer": "#6c757d",
        }
        color = colors.get(obj.role, "#6c757d")
        return format_html(
            '<span style="background:{};color:#fff;padding:2px 8px;border-radius:4px">{}</span>',
            color, obj.role
        )


# ── UploadLog ─────────────────────────────────────────────────────────────────

@admin.register(UploadLog)
class UploadLogAdmin(ObjectFilterMixin, admin.ModelAdmin):
    list_display   = ("id", "filename", "object_name", "doc_type",
                      "status_badge", "uploaded_at", "download_link")
    list_filter    = ("review_status", "doc_type", "object_name")
    search_fields  = ("filename", "object_name", "nas_path")
    readonly_fields = ("uploaded_at", "reviewed_at", "download_link")
    date_hierarchy = None
    actions        = ["action_approve", "action_reject",
                      "action_download_zip", "action_bulk_link"]

    @admin.display(description="Статус")
    def status_badge(self, obj):
        colors = {"pending": "#fd7e14", "approved": "#198754", "rejected": "#dc3545"}
        icons  = {"pending": "⏳", "approved": "✅", "rejected": "❌"}
        color = colors.get(obj.review_status, "#6c757d")
        icon  = icons.get(obj.review_status, "⚪")
        return format_html(
            '<span style="background:{};color:#fff;padding:2px 8px;border-radius:4px">{} {}</span>',
            color, icon, obj.review_status
        )

    @admin.display(description="NAS")
    def download_link(self, obj):
        if obj.nas_path:
            return format_html(
                '<a href="/nas-proxy/?path={}" target="_blank">⬇️ Скачать</a>',
                obj.nas_path
            )
        return "—"

    @admin.action(description="✅ Утвердить выбранные")
    def action_approve(self, request, queryset):
        root = os.path.dirname(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        )
        if root not in sys.path:
            sys.path.insert(0, root)
        import apps.bot.bot_db as db
        from core.services.approvals import approve_doc
        from core.services.notify import notify_doc_approved
        reviewer = request.user.get_full_name() or request.user.username
        ok = err = 0
        for obj in queryset.filter(review_status="pending"):
            result = approve_doc(db, _nas(), obj.id,
                                 request.user.id if request.user.id else 0)
            if result["ok"]:
                ok += 1
                # Sprint 13: notify uploader via Telegram
                try:
                    uploader_row = db.get_upload(obj.id)
                    if uploader_row and uploader_row.get("telegram_id"):
                        notify_doc_approved(
                            telegram_id=int(uploader_row["telegram_id"]),
                            filename=obj.filename,
                            doc_id=obj.id,
                            reviewer=reviewer,
                        )
                except Exception:
                    pass
            else:
                err += 1
        self.message_user(request, f"Утверждено: {ok}, ошибок: {err}")

    @admin.action(description="❌ Отклонить выбранные (с причиной)")
    def action_reject(self, request, queryset):
        """
        Если в POST-запросе есть 'reject_reason' — используем его.
        Иначе — отвечаем HTML-формой с диалогом ввода причины.
        """
        reason = request.POST.get("reject_reason", "").strip()

        # ── Первый вызов: нет причины → показываем диалог ──
        if not reason:
            ids = ",".join(str(obj.pk) for obj in queryset)
            from django.shortcuts import render as _render
            return _render(request, "adminpanel/reject_reason_form.html", {
                "queryset": queryset,
                "ids": ids,
                "action": "action_reject",
                "opts": self.model._meta,
            })

        # ── Второй вызов: причина есть → выполняем отклонение ──
        _ensure_root_in_path()
        import apps.bot.bot_db as db
        from core.services.approvals import reject_doc
        from core.services.notify import notify_doc_rejected
        reviewer = request.user.get_full_name() or request.user.username

        # При возврате из формы queryset может быть пустым — берём по ids
        ids_str = request.POST.get("reject_ids", "")
        if ids_str:
            from .models import UploadLog as _UL
            try:
                id_list = [int(i) for i in ids_str.split(",") if i.strip()]
                queryset = _UL.objects.filter(pk__in=id_list)
            except ValueError:
                pass

        ok = 0
        for obj in queryset.filter(review_status="pending"):
            reject_doc(db, _nas(), obj.id,
                       request.user.id if request.user.id else 0,
                       reason)
            ok += 1
            try:
                uploader_row = db.get_upload(obj.id)
                if uploader_row and uploader_row.get("telegram_id"):
                    notify_doc_rejected(
                        telegram_id=int(uploader_row["telegram_id"]),
                        filename=obj.filename,
                        doc_id=obj.id,
                        reviewer=reviewer,
                        reason=reason,
                    )
            except Exception:
                pass
        self.message_user(request, f"Отклонено: {ok} · Причина: {reason}")

    @admin.action(description="🔗 Создать связи с выбранными (doc_links)")
    def action_bulk_link(self, request, queryset):
        """
        Mass-link all selected UploadLog entries to each other via doc_links.
        Creates N*(N-1)/2 pairs for deduplication/grouping.
        """
        _ensure_root_in_path()
        import apps.bot.bot_db as db
        ids = list(queryset.values_list("id", flat=True))
        created = 0
        for i, a in enumerate(ids):
            for b in ids[i + 1:]:
                result = db.add_link("upload", a, "upload", b,
                                     created_by=request.user.id or 0)
                if result:
                    created += 1
        self.message_user(request, f"Создано связей: {created}")

    @admin.action(description="📦 Скачать выбранные как ZIP")
    def action_download_zip(self, request, queryset):
        import zipfile
        buf = io.BytesIO()
        nas = _nas()
        added = 0
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for obj in queryset[:20]:
                content = nas.download(obj.nas_path)
                if content:
                    zf.writestr(f"{obj.object_name}/{obj.filename}", content)
                    added += 1
        buf.seek(0)
        resp = HttpResponse(buf.read(), content_type="application/zip")
        resp["Content-Disposition"] = 'attachment; filename="documents.zip"'
        return resp


# ── ExpiryItem ────────────────────────────────────────────────────────────────

@admin.register(ExpiryItem)
class ExpiryItemAdmin(ObjectFilterMixin, admin.ModelAdmin):
    list_display  = ("id", "title", "object_name", "expires_at",
                     "days_left", "status")
    list_filter   = ("status", "object_name")
    search_fields = ("title", "object_name")
    readonly_fields = ("created_at",)
    actions       = ["action_archive"]

    @admin.display(description="Осталось дней")
    def days_left(self, obj):
        from datetime import date
        try:
            delta = (date.fromisoformat(obj.expires_at) - date.today()).days
            if delta < 0:
                return format_html('<span style="color:red">просрочено ({}д)</span>', abs(delta))
            elif delta <= 7:
                return format_html('<span style="color:orange">{}д</span>', delta)
            return format_html('<span style="color:green">{}д</span>', delta)
        except Exception:
            return "—"

    @admin.action(description="📦 Архивировать")
    def action_archive(self, request, queryset):
        updated = queryset.update(status="archived")
        self.message_user(request, f"Архивировано: {updated}")


# ── FinanceDoc ────────────────────────────────────────────────────────────────

@admin.register(FinanceDoc)
class FinanceDocAdmin(ObjectFilterMixin, admin.ModelAdmin):
    list_display   = ("id", "filename", "object_name", "doc_type",
                      "amount", "counterparty", "status_badge",
                      "created_at", "download_link")
    list_filter    = ("status", "doc_type", "object_name")
    search_fields  = ("filename", "counterparty", "object_name")
    readonly_fields = ("created_at", "updated_at", "download_link")
    actions        = ["export_csv_action", "mass_set_na_proverke",
                      "mass_set_approved", "mass_set_rejected", "mass_set_paid"]

    @admin.display(description="Статус")
    def status_badge(self, obj):
        colors = {
            "черновик": "#6c757d", "на_проверке": "#0d6efd",
            "утверждён": "#198754", "отклонён": "#dc3545", "оплачен": "#fd7e14",
        }
        color = colors.get(obj.status, "#aaa")
        return format_html(
            '<span style="background:{};color:#fff;padding:2px 8px;border-radius:4px">{}</span>',
            color, obj.status
        )

    @admin.display(description="NAS")
    def download_link(self, obj):
        if obj.nas_path:
            return format_html(
                '<a href="/nas-proxy/?path={}" target="_blank">⬇️ Скачать</a>',
                obj.nas_path
            )
        return "—"

    @admin.action(description="📊 Экспорт CSV")
    def export_csv_action(self, request, queryset):
        out = io.StringIO()
        out.write("\ufeff")
        writer = csv.writer(out, delimiter=";")
        writer.writerow(["ID", "Объект", "Тип", "Файл", "Сумма",
                         "Контрагент", "Статус", "Создан", "Обновлён"])
        for obj in queryset:
            writer.writerow([
                obj.id, obj.object_name, obj.doc_type, obj.filename,
                obj.amount or "", obj.counterparty or "",
                obj.status, obj.created_at, obj.updated_at,
            ])
        resp = HttpResponse(out.getvalue().encode("utf-8-sig"),
                            content_type="text/csv; charset=utf-8-sig")
        resp["Content-Disposition"] = 'attachment; filename="finance.csv"'
        return resp

    # ── Mass finance status actions ─────────────────────────────────────────

    @admin.action(description="🔍 → На проверку (массово)")
    def mass_set_na_proverke(self, request, queryset):
        self._mass_status(request, queryset, "черновик", "на_проверке")

    @admin.action(description="✅ → Утвердить (массово)")
    def mass_set_approved(self, request, queryset):
        self._mass_status(request, queryset, "на_проверке", "утверждён")

    @admin.action(description="❌ → Отклонить (массово)")
    def mass_set_rejected(self, request, queryset):
        self._mass_status(request, queryset, "на_проверке", "отклонён")

    @admin.action(description="💰 → Оплачен (массово)")
    def mass_set_paid(self, request, queryset):
        self._mass_status(request, queryset, "утверждён", "оплачен")

    def _mass_status(self, request, queryset, expected_current: str, new_status: str):
        _ensure_root_in_path()
        import apps.bot.bot_db as db
        changed = 0
        for obj in queryset.filter(status=expected_current):
            db.update_finance_status(obj.id, new_status,
                                     request.user.id or 0, "Массовое изменение через Web")
            changed += 1
        skipped = queryset.exclude(status=expected_current).count()
        msg = f"Изменено: {changed}"
        if skipped:
            msg += f" (пропущено {skipped}: не тот статус)"
        self.message_user(request, msg)


@admin.register(FinanceStatusLog)
class FinanceStatusLogAdmin(admin.ModelAdmin):
    list_display  = ("id", "finance_doc_id", "old_status", "new_status",
                     "changed_by", "comment", "changed_at")
    readonly_fields = ("id", "finance_doc_id", "old_status", "new_status",
                       "changed_by", "comment", "changed_at")
    list_filter   = ("new_status",)

    def has_add_permission(self, request): return False
    def has_change_permission(self, request, obj=None): return False


# ── Problem ───────────────────────────────────────────────────────────────────

@admin.register(Problem)
class ProblemAdmin(admin.ModelAdmin):
    list_display  = ("id", "label", "description_short", "status",
                     "created_by", "created_at")
    list_filter   = ("status", "label")
    search_fields = ("label", "description")
    actions       = ["close_problems"]

    @admin.display(description="Описание")
    def description_short(self, obj):
        return obj.description[:60] + "…" if len(obj.description) > 60 else obj.description

    @admin.action(description="✅ Закрыть выбранные")
    def close_problems(self, request, queryset):
        updated = queryset.update(status="closed")
        self.message_user(request, f"Закрыто: {updated}")


# ── PackageLog ────────────────────────────────────────────────────────────────

@admin.register(PackageLog)
class PackageLogAdmin(admin.ModelAdmin):
    list_display  = ("id", "object_name", "period", "doc_types",
                     "file_count", "status", "created_at", "download_link")
    readonly_fields = ("created_at", "download_link")
    list_filter   = ("status", "object_name")

    @admin.display(description="Скачать")
    def download_link(self, obj):
        if obj.nas_zip_path:
            return format_html(
                '<a href="/nas-proxy/?path={}" target="_blank">⬇️ ZIP</a>',
                obj.nas_zip_path
            )
        return "—"

    def has_add_permission(self, request): return False


# ── Report ────────────────────────────────────────────────────────────────────

@admin.register(Report)
class ReportAdmin(admin.ModelAdmin):
    list_display  = ("id", "object_name", "report_date", "status",
                     "telegram_id", "created_at")
    list_filter   = ("status", "object_name")
    readonly_fields = ("created_at",)

    def has_add_permission(self, request): return False


# ── AuditLog ──────────────────────────────────────────────────────────────────

@admin.register(DocLink)
class DocLinkAdmin(admin.ModelAdmin):
    list_display  = ("id", "from_type", "from_id", "arrow", "to_type", "to_id",
                     "created_by", "created_at")
    list_filter   = ("from_type", "to_type")
    search_fields = ("from_id", "to_id")
    actions       = ["delete_selected_links"]

    @admin.display(description="")
    def arrow(self, obj):
        return format_html('<span style="color:#f39c12;font-weight:bold">→</span>')

    @admin.action(description="🗑 Удалить выбранные связи")
    def delete_selected_links(self, request, queryset):
        count = queryset.count()
        queryset.delete()
        self.message_user(request, f"Удалено связей: {count}")


@admin.register(NasObject)
class NasObjectAdmin(admin.ModelAdmin):
    list_display  = ("id", "name", "nas_path", "is_active", "created_at", "summary_link")
    list_filter   = ("is_active",)
    search_fields = ("name",)
    list_editable = ("is_active",)
    readonly_fields = ("created_at", "summary_link")

    @admin.display(description="Сводка")
    def summary_link(self, obj):
        return format_html(
            '<a href="/objects/{}/" target="_blank">📊 Сводка</a>',
            obj.name
        )


@admin.register(AuditLog)
class AuditLogAdmin(admin.ModelAdmin):
    list_display   = ("id", "telegram_id", "action", "entity_type",
                      "entity_id", "detail_short", "created_at")
    list_filter    = ("action", "entity_type")
    search_fields  = ("action", "entity_type", "detail")
    readonly_fields = ("id", "telegram_id", "action", "entity_type",
                       "entity_id", "detail", "created_at")

    @admin.display(description="Детали")
    def detail_short(self, obj):
        return obj.detail[:80] + "…" if len(obj.detail) > 80 else obj.detail

    def has_add_permission(self, request): return False
    def has_change_permission(self, request, obj=None): return False
    def has_delete_permission(self, request, obj=None): return False


# ── Sprint 11: Document Registry ──────────────────────────────────────────────

@admin.register(Document)
class DocumentAdmin(ObjectFilterMixin, admin.ModelAdmin):
    list_display   = ("id", "original_filename", "object_name", "category_badge",
                      "doc_type", "status_badge", "file_size_fmt",
                      "created_at", "download_link", "dedupe_flag", "card_link")
    list_filter    = ("status", "category", "doc_type", "object_name")
    search_fields  = ("original_filename", "object_name", "nas_path", "file_hash")
    readonly_fields = ("id", "file_hash", "file_size", "created_at", "updated_at",
                       "download_link", "dedupe_flag")
    actions        = ["action_approve_docs", "action_reject_docs",
                      "action_archive_docs", "export_doc_csv",
                      "export_doc_xlsx", "export_doc_pdf"]

    @admin.display(description="Категория")
    def category_badge(self, obj):
        colors = {
            "build": "#0d6efd", "finance": "#198754",
            "safety": "#dc3545", "photo": "#6610f2", "other": "#6c757d",
        }
        labels = {
            "build": "Строй", "finance": "Финансы",
            "safety": "ТБ", "photo": "Фото", "other": "Прочее",
        }
        color = colors.get(obj.category, "#aaa")
        label = labels.get(obj.category, obj.category)
        return format_html(
            '<span style="background:{};color:#fff;padding:2px 8px;border-radius:4px">{}</span>',
            color, label
        )

    @admin.display(description="Статус")
    def status_badge(self, obj):
        colors = {"pending": "#fd7e14", "approved": "#198754",
                  "rejected": "#dc3545", "archived": "#6c757d"}
        icons  = {"pending": "⏳", "approved": "✅",
                  "rejected": "❌", "archived": "📦"}
        color = colors.get(obj.status, "#aaa")
        icon  = icons.get(obj.status, "⚪")
        return format_html(
            '<span style="background:{};color:#fff;padding:2px 8px;border-radius:4px">{} {}</span>',
            color, icon, obj.status
        )

    @admin.display(description="Размер")
    def file_size_fmt(self, obj):
        if not obj.file_size:
            return "—"
        if obj.file_size >= 1_048_576:
            return f"{obj.file_size / 1_048_576:.1f} МБ"
        return f"{obj.file_size / 1024:.0f} КБ"

    @admin.display(description="NAS")
    def download_link(self, obj):
        if obj.nas_path:
            return format_html(
                '<a href="/nas-proxy/?path={}" target="_blank">⬇️ Скачать</a>',
                obj.nas_path
            )
        return "—"

    @admin.display(description="Дубликат?")
    def dedupe_flag(self, obj):
        if not obj.file_hash:
            return "—"
        count = Document.objects.filter(file_hash=obj.file_hash).count()
        if count > 1:
            return format_html('<span style="color:orange">⚠️ {} копий</span>', count)
        return format_html('<span style="color:green">✅</span>')

    @admin.display(description="Карточка")
    def card_link(self, obj):
        return format_html(
            '<a href="/doc/{}/" target="_blank" '
            'style="color:#3498db;text-decoration:none;font-weight:600">🔍 Открыть</a>',
            obj.id
        )

    @admin.action(description="✅ Утвердить выбранные документы")
    def action_approve_docs(self, request, queryset):
        updated = queryset.update(status="approved")
        self.message_user(request, f"Утверждено: {updated}")

    @admin.action(description="❌ Отклонить выбранные документы")
    def action_reject_docs(self, request, queryset):
        updated = queryset.update(status="rejected")
        self.message_user(request, f"Отклонено: {updated}")

    @admin.action(description="📦 Архивировать выбранные документы")
    def action_archive_docs(self, request, queryset):
        updated = queryset.update(status="archived")
        self.message_user(request, f"Архивировано: {updated}")

    @admin.action(description="📊 Экспорт в CSV")
    def export_doc_csv(self, request, queryset):
        out = io.StringIO()
        out.write("\ufeff")
        writer = csv.writer(out, delimiter=";")
        writer.writerow(["ID", "Объект", "Категория", "Тип", "Файл",
                         "Статус", "Хэш", "Размер (байт)", "Создан"])
        for obj in queryset:
            writer.writerow([
                obj.id, obj.object_name, obj.category, obj.doc_type or "",
                obj.original_filename, obj.status,
                obj.file_hash or "", obj.file_size or 0, obj.created_at,
            ])
        resp = HttpResponse(out.getvalue().encode("utf-8-sig"),
                            content_type="text/csv; charset=utf-8-sig")
        resp["Content-Disposition"] = 'attachment; filename="documents_registry.csv"'
        return resp

    @admin.action(description="📗 Экспорт в Excel (.xlsx)")
    def export_doc_xlsx(self, request, queryset):
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side,
        )
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Реестр документов"

        # ── Header style ──
        hdr_font  = Font(bold=True, color="FFFFFF", size=10)
        hdr_fill  = PatternFill("solid", fgColor="3D5166")
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_side = Side(style="thin", color="CCCCCC")
        cell_border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side,
        )

        headers = [
            "ID", "Объект", "Категория", "Тип документа", "Имя файла",
            "Статус", "Размер (KB)", "SHA-256", "Создан",
        ]
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font  = hdr_font
            cell.fill  = hdr_fill
            cell.alignment = hdr_align
            cell.border = cell_border

        # ── Data rows ──
        status_colors = {
            "pending":  "FFF3CD",
            "approved": "D4EDDA",
            "rejected": "F8D7DA",
            "archived": "E9ECEF",
        }
        cat_labels = dict(CATEGORY_CHOICES) if hasattr(self, '_cat') else {
            "build": "Строительство", "finance": "Финансы",
            "safety": "ТБ/ОТ", "photo": "Фото", "other": "Прочее",
        }

        for row_idx, obj in enumerate(queryset, start=2):
            size_kb = round((obj.file_size or 0) / 1024, 1) if obj.file_size else ""
            row_data = [
                obj.id,
                obj.object_name,
                cat_labels.get(obj.category, obj.category),
                obj.doc_type or "",
                obj.original_filename,
                obj.status,
                size_kb,
                (obj.file_hash or "")[:16] + "…" if obj.file_hash else "",
                (obj.created_at or "")[:16],
            ]
            ws.append(row_data)
            fill_color = status_colors.get(obj.status, "FFFFFF")
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill   = PatternFill("solid", fgColor=fill_color)
                cell.border = cell_border
                cell.alignment = Alignment(vertical="center", wrap_text=False)

        # ── Column widths ──
        col_widths = [6, 22, 14, 20, 36, 12, 10, 20, 17]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── Freeze header ──
        ws.freeze_panes = "A2"

        # ── Legend sheet ──
        ws_legend = wb.create_sheet("Легенда")
        ws_legend.append(["Статус", "Значение", "Цвет"])
        legend_data = [
            ("pending",  "На проверке", "FFF3CD"),
            ("approved", "Утверждён",   "D4EDDA"),
            ("rejected", "Отклонён",    "F8D7DA"),
            ("archived", "Архив",       "E9ECEF"),
        ]
        for st, label, color in legend_data:
            ws_legend.append([st, label])
            ws_legend.cell(row=ws_legend.max_row, column=1).fill = PatternFill("solid", fgColor=color)

        # ── Save to response ──
        import io as _io
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        from datetime import date
        today = date.today().strftime("%Y%m%d")
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="documents_{today}.xlsx"'
        return resp


# ── Sprint 9.1+: UserObject (permissions by object) ───────────────────────────

@admin.register(UserObject)
class UserObjectAdmin(admin.ModelAdmin):
    list_display  = ("id", "telegram_id", "object_name", "granted_by", "granted_at")
    list_filter   = ("object_name",)
    search_fields = ("telegram_id", "object_name")
    actions       = ["revoke_access"]

    @admin.action(description="📥 Экспорт в PDF")
    def export_doc_pdf(self, request, queryset):
        _ensure_root_in_path()
        from core.services.pdf_report import build_registry_pdf
        from datetime import date
        try:
            pdf_bytes = build_registry_pdf(list(queryset[:500]))
        except Exception as exc:
            self.message_user(request, f"Ошибка PDF: {exc}", level="error")
            return
        fname = f"registry_{date.today().strftime('%Y%m%d')}.pdf"
        resp = HttpResponse(pdf_bytes, content_type="application/pdf")
        resp["Content-Disposition"] = f'attachment; filename="{fname}"'
        return resp


    @admin.action(description="🚫 Отозвать доступ к объекту")
    def revoke_access(self, request, queryset):
        count = queryset.count()
        queryset.delete()
        self.message_user(request, f"Доступ отозван у {count} записей")


# ── Sprint 12: OCR Results — Human-in-the-loop ───────────────────────────────

@admin.register(OcrResult)
class OcrResultAdmin(admin.ModelAdmin):
    list_display  = (
        "id", "status_badge", "upload_link", "doc_number", "doc_date",
        "expires_at", "counterparty", "amount_fmt", "confidence_bar", "created_at",
    )
    list_filter   = ("status",)
    search_fields = ("doc_number", "counterparty", "raw_text")
    readonly_fields = (
        "upload_id", "doc_id", "raw_text_preview",
        "confidence", "created_at", "reviewed_by", "reviewed_at",
    )
    fieldsets = (
        ("🔍 Распознанные поля (редактируйте и подтвердите)", {
            "fields": ("doc_number", "doc_date", "expires_at", "counterparty", "amount"),
        }),
        ("📊 Системная информация", {
            "fields": ("upload_id", "doc_id", "confidence", "status",
                       "reviewed_by", "reviewed_at", "created_at"),
            "classes": ("collapse",),
        }),
        ("📄 Распознанный текст", {
            "fields": ("raw_text_preview",),
            "classes": ("collapse",),
        }),
    )
    actions = [
        "action_confirm_ocr",
        "action_reject_ocr",
        "action_create_expiry",
    ]

    def status_badge(self, obj):
        colors = {"pending": "#e67e22", "confirmed": "#27ae60", "rejected": "#e74c3c"}
        labels = {"pending": "⏳ Ожидает", "confirmed": "✅ Подтверждено", "rejected": "❌ Отклонено"}
        color = colors.get(obj.status, "#95a5a6")
        label = labels.get(obj.status, obj.status)
        return format_html(
            '<span style="background:{};color:#fff;padding:3px 10px;'
            'border-radius:12px;font-size:.8rem;font-weight:600">{}</span>',
            color, label,
        )
    status_badge.short_description = "Статус"

    def upload_link(self, obj):
        if obj.upload_id:
            return format_html(
                '<a href="/admin/adminpanel/uploadlog/{}/change/">📄 #{}</a>',
                obj.upload_id, obj.upload_id,
            )
        return "—"
    upload_link.short_description = "Загрузка"

    def amount_fmt(self, obj):
        if obj.amount is not None:
            return f"{obj.amount:,.2f}"
        return "—"
    amount_fmt.short_description = "Сумма"

    def confidence_bar(self, obj):
        pct = obj.confidence or 0
        color = "#27ae60" if pct >= 75 else "#e67e22" if pct >= 25 else "#e74c3c"
        return format_html(
            '<div style="background:#eee;border-radius:4px;width:80px;height:10px">'
            '<div style="background:{};width:{}%;height:100%;border-radius:4px"></div>'
            '</div> {}%',
            color, pct, pct,
        )
    confidence_bar.short_description = "Точность"

    def raw_text_preview(self, obj):
        if obj.raw_text:
            escaped = obj.raw_text[:1500].replace("<", "&lt;").replace(">", "&gt;")
            return format_html(
                '<pre style="white-space:pre-wrap;font-size:.8rem;'
                'max-height:300px;overflow:auto;background:#f8f9fa;'
                'padding:10px;border-radius:6px">{}</pre>',
                escaped,
            )
        return "Текст не извлечён"
    raw_text_preview.short_description = "Извлечённый текст"

    @admin.action(description="✅ Подтвердить распознанные данные")
    def action_confirm_ocr(self, request, queryset):
        _ensure_root_in_path()
        import apps.bot.bot_db as db
        from core.services.notify import notify_async
        confirmed = 0
        for obj in queryset.filter(status="pending"):
            data = {
                "doc_number":   obj.doc_number,
                "doc_date":     obj.doc_date,
                "expires_at":   obj.expires_at,
                "counterparty": obj.counterparty,
                "amount":       obj.amount,
            }
            db.confirm_ocr_result(obj.id, data, reviewed_by=0)
            # Auto-create expiry item if expiry date found
            if obj.expires_at and obj.upload_id:
                try:
                    upload = db.get_upload(obj.upload_id)
                    if upload:
                        db.add_expiry(
                            object_name=upload.get("object_name", ""),
                            doc_type=upload.get("doc_type", "Документ"),
                            expires_at=obj.expires_at,
                            upload_id=obj.upload_id,
                            created_by=0,
                        )
                except Exception:
                    pass
            confirmed += 1
        self.message_user(request, f"Подтверждено: {confirmed} результатов OCR")

    @admin.action(description="❌ Отклонить результат OCR")
    def action_reject_ocr(self, request, queryset):
        _ensure_root_in_path()
        import apps.bot.bot_db as db
        for obj in queryset.filter(status="pending"):
            db.reject_ocr_result(obj.id, reviewed_by=0)
        self.message_user(request, f"Отклонено: {queryset.count()} результатов")

    @admin.action(description="📅 Создать срок из expires_at")
    def action_create_expiry(self, request, queryset):
        _ensure_root_in_path()
        import apps.bot.bot_db as db
        created = 0
        for obj in queryset:
            if not obj.expires_at:
                continue
            try:
                upload = db.get_upload(obj.upload_id) if obj.upload_id else None
                obj_name = (upload or {}).get("object_name", "Неизвестный объект")
                doc_name = (upload or {}).get("doc_type", "Документ")
                db.add_expiry(
                    object_name=obj_name,
                    doc_type=doc_name,
                    expires_at=obj.expires_at,
                    upload_id=obj.upload_id,
                    created_by=0,
                )
                created += 1
            except Exception as e:
                self.message_user(request, f"Ошибка OCR #{obj.id}: {e}", level="error")
        self.message_user(request, f"Создано сроков: {created}")
